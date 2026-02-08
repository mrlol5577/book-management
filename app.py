import os
from flask import Flask, render_template, url_for, request, redirect, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from flask_migrate import Migrate
import json
import io
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

# Створюємо папку instance
basedir = os.path.abspath(os.path.dirname(__file__))
instance_path = os.path.join(basedir, 'instance')
if not os.path.exists(instance_path):
    os.makedirs(instance_path)

app = Flask(__name__)

# ⚙️ УНІВЕРСАЛЬНЕ НАЛАШТУВАННЯ БД
# Працює і локально (SQLite), і на сервері (PostgreSQL)
database_url = os.environ.get("DATABASE_URL")

if database_url:
    # На сервері - використовуй PostgreSQL
    # Виправляємо postgres:// на postgresql:// для SQLAlchemy
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    print("🌐 Використовується PostgreSQL (сервер)")
else:
    # Локально - використовуй SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///library.db'
    print("💻 Використовується SQLite (локально)")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your-secret-key-here-change-this'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Будь ласка, увійдіть для доступу до цієї сторінки.'

# Моделі
class Reader(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    surname = db.Column(db.String(200), nullable=False)
    phone = db.Column(db.String(50), nullable=False)

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False) 
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='admin')  

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Book(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name_book = db.Column(db.String(500), nullable=False)  # ✅ Збільшено з 100 до 500
    author = db.Column(db.String(500), nullable=False)     # ✅ Збільшено з 100 до 500
    surname = db.Column(db.String(200), default='')        # ✅ Збільшено з 100 до 200
    ean = db.Column(db.Text, default='-')
    buyer = db.Column(db.String(200), nullable=False)      # ✅ Збільшено з 100 до 200
    phone = db.Column(db.String(50), nullable=False)       # ✅ Збільшено з 20 до 50
    stat = db.Column(db.String(20), nullable=False)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    enddate = db.Column(db.DateTime, default=datetime.utcnow)
    history = db.Column(db.Text, default='')               # ✅ Змінено з String(100) на Text

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ====== ІМПОРТ З EXCEL ======
@app.route('/import-excel', methods=['GET', 'POST'])
@login_required
def import_excel():
    if current_user.role != 'superadmin':
        flash('❌ У вас немає прав для імпорту даних!', 'danger')
        return redirect('/books')
    
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('❌ Файл не вибрано!', 'danger')
            return redirect(request.url)
        
        file = request.files['excel_file']
        
        if file.filename == '':
            flash('❌ Файл не вибрано!', 'danger')
            return redirect(request.url)
        
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                # Читаємо Excel файл з openpyxl
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                
                stats = {
                    'added': 0,
                    'errors': []
                }
                
                # Отримуємо заголовки (перший рядок)
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).lower().strip() if cell.value else '')
                
                # Знаходимо індекси потрібних колонок
                column_mapping = {}
                for idx, header in enumerate(headers):
                    # Варіанти для назви книги
                    if header in ['name_book', 'назва', 'книга', 'name', 'название']:
                        column_mapping['name_book'] = idx
                    
                    # Варіанти для автора
                    elif header in ['author', 'автор', 'writer']:
                        column_mapping['author'] = idx
                    
                    # Варіанти для EAN
                    elif header in ['ean', 'isbn', 'код', 'code']:
                        column_mapping['ean'] = idx
                
                # Перевіряємо чи є обов'язкові колонки
                if 'name_book' not in column_mapping or 'author' not in column_mapping:
                    flash(f'❌ У файлі відсутні обов\'язкові колонки! Потрібні: "name_book" (або "назва") та "author" (або "автор"). Знайдено колонки: {", ".join(headers)}', 'danger')
                    return redirect(request.url)
                
                # Обробляємо кожен рядок (починаючи з 2-го)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Отримуємо значення
                        name_book = str(row[column_mapping['name_book']]).strip() if row[column_mapping['name_book']] else ''
                        author = str(row[column_mapping['author']]).strip() if row[column_mapping['author']] else ''
                        
                        # Перевіряємо чи не пусті
                        if not name_book or name_book == 'None':
                            stats['errors'].append(f'Рядок {row_idx}: Відсутня назва книги')
                            continue
                        
                        if not author or author == 'None':
                            stats['errors'].append(f'Рядок {row_idx}: Відсутній автор')
                            continue
                        
                        # Отримуємо EAN (якщо є)
                        if 'ean' in column_mapping:
                            ean = str(row[column_mapping['ean']]).strip() if row[column_mapping['ean']] else '-'
                            if ean == 'None' or not ean:
                                ean = '-'
                        else:
                            ean = '-'
                        
                        # Створюємо нову книгу
                        book = Book(
                            name_book=name_book,
                            author=author,
                            ean=ean,
                            buyer='',
                            phone='',
                            stat='доступна',
                            date=datetime.utcnow(),
                            enddate=datetime.utcnow(),
                            history=''
                        )
                        
                        db.session.add(book)
                        stats['added'] += 1
                        
                    except Exception as e:
                        stats['errors'].append(f'Рядок {row_idx}: {str(e)}')
                        continue
                
                # Зберігаємо всі зміни
                db.session.commit()
                
                # Повідомлення про результат
                message = f"✅ Імпорт завершено! Додано книг: {stats['added']}"
                if stats['errors']:
                    message += f"\n⚠️ Помилок: {len(stats['errors'])}"
                    # Показуємо перші 5 помилок
                    for error in stats['errors'][:5]:
                        message += f"\n• {error}"
                    if len(stats['errors']) > 5:
                        message += f"\n• ... та ще {len(stats['errors']) - 5} помилок"
                
                flash(message, 'success' if not stats['errors'] else 'warning')
                return redirect('/books')
                
            except Exception as e:
                db.session.rollback()
                flash(f'❌ Помилка при імпорті: {str(e)}', 'danger')
                return redirect(request.url)
        else:
            flash('❌ Невірний формат файлу! Потрібен файл .xlsx або .xls', 'danger')
            return redirect(request.url)
    
    return render_template('import_excel.html')

# ====== ЕКСПОРТ БАЗИ У JSON (працює і локально, і на сервері) ======
@app.route('/download-db-secret-12345')
@login_required
def download_database():
    if current_user.role != 'superadmin': 
        flash('❌ У вас немає прав для завантаження бази даних!', 'danger') 
        return redirect('/books')
    
    try:
        backup_data = {
            'timestamp': datetime.now().isoformat(),
            'books': [],
            'readers': [],
            'users': []
        }
        
        # Експортуємо книги
        for book in Book.query.all():
            backup_data['books'].append({
                'id': book.id,
                'name_book': book.name_book,
                'author': book.author,
                'surname': book.surname,
                'ean': book.ean,
                'buyer': book.buyer,
                'phone': book.phone,
                'stat': book.stat,
                'date': book.date.isoformat() if book.date else None,
                'enddate': book.enddate.isoformat() if book.enddate else None,
                'history': book.history
            })
        
        # Експортуємо читачів
        for reader in Reader.query.all():
            backup_data['readers'].append({
                'id': reader.id,
                'name': reader.name,
                'surname': reader.surname,
                'phone': reader.phone
            })
        
        # Експортуємо користувачів
        for user in User.query.all():
            backup_data['users'].append({
                'id': user.id,
                'username': user.username,
                'password_hash': user.password_hash,
                'role': user.role
            })
        
        # Створюємо JSON файл
        json_data = json.dumps(backup_data, ensure_ascii=False, indent=2)
        buffer = io.BytesIO()
        buffer.write(json_data.encode('utf-8'))
        buffer.seek(0)
        
        filename = f'library_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        
        print(f"📥 Експортовано: {len(backup_data['books'])} книг, {len(backup_data['readers'])} читачів, {len(backup_data['users'])} користувачів")
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/json'
        )
        
    except Exception as e:
        flash(f'❌ Помилка при завантаженні: {str(e)}', 'danger')
        return redirect('/books')

# ====== ІМПОРТ БАЗИ З JSON (працює і локально, і на сервері) ======
@app.route('/restore-db-secret-54321', methods=['GET', 'POST'])
@login_required
def restore_database():
    if current_user.role != 'superadmin':
        flash('❌ У вас немає прав для відновлення бази даних!', 'danger')
        return redirect('/books')
    
    if request.method == 'POST':
        if 'database' not in request.files:
            flash('❌ Файл не вибрано!', 'danger')
            return redirect(request.url)
        
        file = request.files['database']
        
        if file.filename == '':
            flash('❌ Файл не вибрано!', 'danger')
            return redirect(request.url)
        
        # Перевіряємо чи треба очистити базу перед імпортом
        clear_before_import = request.form.get('clear_db') == 'yes'
        
        if file and file.filename.endswith('.json'):
            try:
                # Читаємо JSON
                json_data = file.read().decode('utf-8')
                backup_data = json.loads(json_data)
                
                stats = {
                    'books_restored': 0,
                    'readers_restored': 0,
                    'users_restored': 0,
                    'errors': []
                }
                
                # Очищуємо базу якщо потрібно
                if clear_before_import:
                    try:
                        # Видаляємо всі книги та читачів
                        Book.query.delete()
                        Reader.query.delete()
                        
                        # Видаляємо всіх користувачів КРІМ поточного
                        User.query.filter(User.id != current_user.id).delete()
                        
                        db.session.commit()
                        print("🗑️ База даних очищена перед імпортом")
                    except Exception as e:
                        db.session.rollback()
                        flash(f'❌ Помилка при очищенні бази: {str(e)}', 'danger')
                        return redirect(request.url)
                
                # Відновлюємо книги
                if 'books' in backup_data:
                    for book_data in backup_data['books']:
                        try:
                            book = Book(
                                id=book_data.get('id'),
                                name_book=book_data.get('name_book', ''),
                                author=book_data.get('author', ''),
                                surname=book_data.get('surname', ''),
                                ean=book_data.get('ean', ''),
                                buyer=book_data.get('buyer', ''),
                                phone=book_data.get('phone', ''),
                                stat=book_data.get('stat', 'доступна'),
                                date=datetime.fromisoformat(book_data['date']) if book_data.get('date') else datetime.utcnow(),
                                enddate=datetime.fromisoformat(book_data['enddate']) if book_data.get('enddate') else datetime.utcnow(),
                                history=book_data.get('history', '')
                            )
                            db.session.merge(book)
                            stats['books_restored'] += 1
                        except Exception as e:
                            stats['errors'].append(f"Книга {book_data.get('id')}: {str(e)}")
                
                # Відновлюємо читачів
                if 'readers' in backup_data:
                    for reader_data in backup_data['readers']:
                        try:
                            reader = Reader(
                                id=reader_data.get('id'),
                                name=reader_data.get('name', ''),
                                surname=reader_data.get('surname', ''),
                                phone=reader_data.get('phone', '')
                            )
                            db.session.merge(reader)
                            stats['readers_restored'] += 1
                        except Exception as e:
                            stats['errors'].append(f"Читач {reader_data.get('id')}: {str(e)}")
                
                # Відновлюємо користувачів
                if 'users' in backup_data:
                    for user_data in backup_data['users']:
                        try:
                            if user_data.get('id') != current_user.id:
                                # Перевіряємо чи існує користувач з таким username
                                existing_user = User.query.filter_by(username=user_data.get('username')).first()
                                
                                if existing_user:
                                    # Оновлюємо існуючого користувача
                                    existing_user.password_hash = user_data.get('password_hash', '')
                                    existing_user.role = user_data.get('role', 'admin')
                                    stats['users_restored'] += 1
                                else:
                                    # Створюємо нового користувача
                                    user = User(
                                        id=user_data.get('id'),
                                        username=user_data.get('username', ''),
                                        password_hash=user_data.get('password_hash', ''),
                                        role=user_data.get('role', 'admin')
                                    )
                                    db.session.merge(user)
                                    stats['users_restored'] += 1
                        except Exception as e:
                            stats['errors'].append(f"Користувач {user_data.get('id')}: {str(e)}")
                
                # Комітимо всі зміни
                db.session.commit()
                
                print(f"📤 Імпортовано: {stats['books_restored']} книг, {stats['readers_restored']} читачів, {stats['users_restored']} користувачів")
                
                # Виправляємо sequences тільки для PostgreSQL
                if 'postgresql' in app.config['SQLALCHEMY_DATABASE_URI']:
                    try:
                        db.session.execute(db.text("""
                            SELECT setval(pg_get_serial_sequence('book', 'id'), 
                                   COALESCE((SELECT MAX(id) FROM book), 0) + 1, false);
                        """))
                        db.session.execute(db.text("""
                            SELECT setval(pg_get_serial_sequence('reader', 'id'), 
                                   COALESCE((SELECT MAX(id) FROM reader), 0) + 1, false);
                        """))
                        db.session.execute(db.text("""
                            SELECT setval(pg_get_serial_sequence('user', 'id'), 
                                   COALESCE((SELECT MAX(id) FROM "user"), 0) + 1, false);
                        """))
                        db.session.commit()
                        print("✅ PostgreSQL sequences виправлено автоматично")
                    except Exception as e:
                        print(f"⚠️ Помилка виправлення sequences: {str(e)}")
                
                # Повідомлення про результат
                action = "повністю замінено" if clear_before_import else "оновлено"
                message = f"✅ База даних {action}! Відновлено: Книг: {stats['books_restored']}, Читачів: {stats['readers_restored']}, Користувачів: {stats['users_restored']}"
                if stats['errors']:
                    message += f"\n⚠️ Помилки: {len(stats['errors'])}"
                
                flash(message, 'success')
                return redirect('/books')
                
            except json.JSONDecodeError:
                flash('❌ Невірний формат JSON файлу!', 'danger')
                return redirect(request.url)
            except Exception as e:
                db.session.rollback()
                flash(f'❌ Помилка при відновленні: {str(e)}', 'danger')
                return redirect(request.url)
        else:
            flash('❌ Невірний формат файлу! Потрібен файл .json', 'danger')
            return redirect(request.url)
    
    return render_template('restore_db.html')

# ====== ВИПРАВЛЕННЯ SEQUENCES (тільки для PostgreSQL) ======
@app.route('/fix-sequences-secret-88888')
@login_required
def fix_sequences():
    if current_user.role != 'superadmin':
        flash('❌ У вас немає прав!', 'danger')
        return redirect('/books')
    
    # Перевірка чи це PostgreSQL
    if 'postgresql' not in app.config['SQLALCHEMY_DATABASE_URI']:
        flash('⚠️ Ця функція тільки для PostgreSQL. SQLite не потребує виправлення sequences.', 'info')
        return redirect('/books')
    
    try:
        db.session.execute(db.text("""
            SELECT setval(pg_get_serial_sequence('book', 'id'), 
                   COALESCE((SELECT MAX(id) FROM book), 0) + 1, false);
        """))
        
        db.session.execute(db.text("""
            SELECT setval(pg_get_serial_sequence('reader', 'id'), 
                   COALESCE((SELECT MAX(id) FROM reader), 0) + 1, false);
        """))
        
        db.session.execute(db.text("""
            SELECT setval(pg_get_serial_sequence('user', 'id'), 
                   COALESCE((SELECT MAX(id) FROM "user"), 0) + 1, false);
        """))
        
        db.session.commit()
        
        flash('✅ Послідовності ID успішно виправлено!', 'success')
        return redirect('/books')
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Помилка: {str(e)}', 'danger')
        return redirect('/books')




# ====== ОСНОВНІ МАРШРУТИ ======
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('books'))

    if request.method == 'POST':
        password = request.form['password']
        user_found = None
        for user in User.query.all():
            if user.check_password(password):
                user_found = user
                break

        if user_found:
            login_user(user_found)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('books'))
        else:
            flash('Неправильний пароль', 'danger')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('books'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('books'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if User.query.filter_by(username=username).first():
            flash('Користувач з таким іменем вже існує', 'danger')
            return render_template('register.html')
        
        user = User(username=username)
        user.set_password(password)
        
        try:
            db.session.add(user)
            db.session.commit()
            flash('Реєстрація успішна! Тепер ви можете увійти.', 'success')
            return redirect(url_for('login'))
        except:
            flash('Помилка при реєстрації', 'danger')
    
    return render_template('register.html')

@app.route('/booked')
def booked():
    search_query = request.args.get('search', '')
    if search_query:
        search_query_lower = search_query.lower()
        booked = []
        for book in Book.query.filter(Book.stat == 'видана').all():
            if (search_query_lower in book.name_book.lower() or
                search_query_lower in book.author.lower() or
                search_query_lower in book.ean.lower()):
                booked.append(book)
    else:
        booked = Book.query.filter(Book.stat == 'видана').all()
    return render_template('booked.html', booked=booked, search_query=search_query)

@app.route('/notbook')
def notbook():
    search_query = request.args.get('search', '')
    if search_query:
        search_query_lower = search_query.lower()
        notbook = []
        for book in Book.query.filter(Book.stat == 'доступна').all():
            if (search_query_lower in book.name_book.lower() or
                search_query_lower in book.author.lower() or
                search_query_lower in book.ean.lower()):
                notbook.append(book)
    else:
        notbook = Book.query.filter(Book.stat == 'доступна').all()
    return render_template('notbook.html', notbook=notbook, search_query=search_query)

@app.route('/')
@app.route('/books')
def books():
    search_query = request.args.get('search', '')
    if search_query:
        search_query_lower = search_query.lower()
        books = []
        for book in Book.query.all():
            if (search_query_lower in book.name_book.lower() or
                search_query_lower in book.author.lower() or
                search_query_lower in book.ean.lower()):
                books.append(book)
    else:
        books = Book.query.all()
    return render_template('value_books.html', books=books, search_query=search_query)

@app.route('/readers')
@login_required
def readers():
    search_query = request.args.get('search', '')
    if search_query:
        search_query_lower = search_query.lower()
        readers = []
        for reader in Reader.query.all():
            if (search_query_lower in reader.name.lower() or
                search_query_lower in reader.surname.lower() or
                search_query_lower in reader.phone.lower()):
                readers.append(reader)
    else:
        readers = Reader.query.all()
    
    # Сортуємо читачів за прізвищем, потім за іменем
    readers = sorted(readers, key=lambda r: (r.surname.lower(), r.name.lower()))
    
    # Перевіряємо чи є у читача книги
    readers_with_books = []
    for reader in readers:
        # Шукаємо видані книги з телефоном читача
        has_books = Book.query.filter(
            Book.stat == 'видана',
            Book.phone == reader.phone
        ).count() > 0
        readers_with_books.append({
            'reader': reader,
            'has_books': has_books
        })
    
    return render_template('readers.html', readers_data=readers_with_books, search_query=search_query, total_readers=len(readers))

@app.route('/books/<int:id>', methods=['POST', 'GET'])
@login_required
def change(id):
    book = Book.query.get(id)
    if request.method == 'POST':
        enddate_str = request.form.get('enddate')
        if enddate_str:
            enddate = datetime.strptime(enddate_str, '%Y-%m-%d')
        else:
            enddate = datetime.utcnow()
        
        if book.buyer and book.buyer.strip():
            old_buyer = book.buyer
            old_phone = book.phone if book.phone else 'Немає'
            start_date = book.date.strftime('%d.%m.%Y') if book.date else 'Немає'
            end_date_formatted = enddate.strftime('%d.%m.%Y')
            new_history_entry = f"{old_buyer} ({old_phone}) - з {start_date} до {end_date_formatted}"
            if book.history:
                book.history = new_history_entry + " | " + book.history
            else:
                book.history = new_history_entry
        
        new_stat = request.form['stat']
        if new_stat == 'видана':
            buyer = request.form.get('buyer', '').strip()
            phone = request.form.get('phone', '').strip()
            surname = request.form.get('surname', '').strip()
            if not buyer or not phone or not surname:
                flash('⚠️ Заповніть всі поля: ім\'я, прізвище та телефон!', 'warning')
                return render_template('change.html', book=book)
            book.buyer = buyer
            book.phone = phone
            book.surname = surname
            book.stat = 'видана'
            book.date = datetime.utcnow()
            book.enddate = enddate
        else:
            book.buyer = ''
            book.phone = ''
            book.surname = ''
            book.stat = 'доступна'
            book.enddate = enddate
            book.date = datetime.utcnow()

        try:
            db.session.commit()
            if new_stat == 'видана':
                existing_reader = Reader.query.filter_by(phone=phone).first()
                if not existing_reader:
                    new_reader = Reader(name=buyer, surname=surname, phone=phone)
                    db.session.add(new_reader)
                    db.session.commit()
            flash('✅ Дані успішно оновлено!', 'success')
            return redirect('/books')
        except Exception as e:
            db.session.rollback()
            flash(f'⚠️ Помилка: {str(e)}', 'danger')
            return render_template('change.html', book=book)
    return render_template('change.html', book=book)

@app.route('/books/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_book(id):
    book = Book.query.get_or_404(id)
    if request.method == 'POST':
        name_book = request.form.get('name_book', '').strip()
        author = request.form.get('author', '').strip()
        ean = request.form.get('ean', '').strip()
        if not name_book or not author:
            flash('⚠️ Назва книги та автор - обов\'язкові поля!', 'warning')
            return render_template('edit_book.html', book=book)
        book.name_book = name_book
        book.author = author
        book.ean = ean
        try:
            db.session.commit()
            flash('✅ Книгу успішно оновлено!', 'success')
            return redirect(f'/books/{book.id}')
        except Exception as e:
            db.session.rollback()
            flash(f'⚠️ Помилка при оновленні: {str(e)}', 'danger')
            return render_template('edit_book.html', book=book)
    return render_template('edit_book.html', book=book)

@app.route('/search_books')
def search_books():
    q = request.args.get('q', '').lower()
    if not q or len(q) < 1:
        return jsonify({'results': []})
    
    # Шукаємо книги
    books = []
    for book in Book.query.all():
        if (q in book.name_book.lower() or 
            q in book.author.lower() or 
            q in book.ean.lower()):
            books.append(book)
    
    # Формуємо результати
    results = []
    for book in books[:10]:  # Показуємо перші 10 результатів
        results.append({
            'id': book.id,
            'name_book': book.name_book,
            'author': book.author,
            'ean': book.ean,
            'stat': book.stat
        })
    
    return jsonify({'results': results})

@app.route('/search_authors')
def search_authors():
    q = request.args.get('q', '').lower()
    if not q or len(q) < 1:
        return jsonify({'results': []})
    
    # Шукаємо унікальних авторів
    authors = db.session.query(Book.author).distinct().all()
    
    # Фільтруємо авторів за запитом
    matching_authors = []
    for author_tuple in authors:
        author = author_tuple[0]
        if q in author.lower():
            matching_authors.append(author)
    
    # Сортуємо за релевантністю (автори що починаються з запиту йдуть першими)
    matching_authors.sort(key=lambda x: (not x.lower().startswith(q), x.lower()))
    
    # Повертаємо перші 10 результатів
    results = [{'author': author} for author in matching_authors[:10]]
    
    return jsonify({'results': results})

@app.route('/create', methods=['POST', 'GET'])
@login_required
def create():
    if request.method == 'POST':
        name_book = request.form['name_book']
        author = request.form['author']
        ean = request.form['ean']
        buyer = request.form['buyer']
        phone = request.form['phone']   
        stat = request.form['stat']   
        date_str = request.form.get('date')
        if date_str:
            date = datetime.strptime(date_str, '%Y-%m-%d')
        else:
            date = datetime.utcnow()
        books = Book(name_book=name_book, author=author, ean=ean, buyer=buyer, phone=phone, stat=stat, date=date)
        try:
            db.session.add(books)
            db.session.commit()
            flash('Книгу успішно додано!', 'success')
            return redirect('/books')
        except Exception as e:
            flash(f'При добавленні статті сталася помилка: {str(e)}', 'danger')
    return render_template('create.html')

@app.route('/rules')
def rules():
    return render_template('rules.html')

@app.route('/reg', methods=['POST', 'GET'])
def reg():
    if request.method == 'POST':
        name = request.form['name']
        surname = request.form['surname']
        phone = request.form['phone']
        reader = Reader(name=name,surname=surname,phone=phone)
        try:
            db.session.add(reader)
            db.session.commit()
            return redirect('/books')
        except Exception as e:
            flash(f'При добавленні статті сталася помилка: {str(e)}', 'danger')
    return render_template('reg.html')

@app.route('/books/<int:id>/del')
@login_required
def post_delete(id):
    if current_user.role != 'superadmin':
        flash('❌ У вас немає прав на видалення!', 'danger')
        return redirect('/books')
    book = Book.query.get_or_404(id)
    try:
        db.session.delete(book)
        db.session.commit()
        flash('✅ Книгу видалено!', 'success')
    except:
        flash('❌ Помилка при видаленні', 'danger')
    return redirect('/books')

@app.route('/readers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_reader(id):
    reader = Reader.query.get_or_404(id)
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        surname = request.form.get('surname', '').strip()
        phone = request.form.get('phone', '').strip()
        
        if not name or not surname or not phone:
            flash('⚠️ Всі поля обов\'язкові!', 'warning')
            return render_template('edit_reader.html', reader=reader)
        
        # Перевіряємо чи не використовується вже такий телефон іншим читачем
        existing_reader = Reader.query.filter(Reader.phone == phone, Reader.id != id).first()
        if existing_reader:
            flash('⚠️ Читач з таким телефоном вже існує!', 'warning')
            return render_template('edit_reader.html', reader=reader)
        
        old_phone = reader.phone
        reader.name = name
        reader.surname = surname
        reader.phone = phone
        
        try:
            # Оновлюємо телефон у всіх книгах цього читача
            if old_phone != phone:
                books = Book.query.filter_by(phone=old_phone).all()
                for book in books:
                    book.phone = phone
            
            db.session.commit()
            flash('✅ Читача успішно оновлено!', 'success')
            return redirect('/readers')
        except Exception as e:
            db.session.rollback()
            flash(f'⚠️ Помилка при оновленні: {str(e)}', 'danger')
            return render_template('edit_reader.html', reader=reader)
    
    return render_template('edit_reader.html', reader=reader)

@app.route('/readers/<int:id>/del')
@login_required
def reader_delete(id):
    if current_user.role not in ['admin', 'superadmin']:
        flash('❌ У вас немає прав на видалення!', 'danger')
        return redirect('/readers')
    
    reader = Reader.query.get_or_404(id)
    
    try:
        db.session.delete(reader)
        db.session.commit()
        flash('✅ Читача видалено!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Помилка при видаленні: {str(e)}', 'danger')
    
    return redirect('/readers')
    
@app.route('/search_reader')
@login_required
def search_reader():
    q = request.args.get('q', '')
    if not q:
        return {'results': []}
    readers = Reader.query.filter(Reader.name.ilike(f'%{q}%')).limit(5).all()
    results = []
    for r in readers:
        results.append({'name': r.name, 'surname': r.surname, 'phone': r.phone})
    return {'results': results}

# ====== ІНІЦІАЛІЗАЦІЯ БД ======
with app.app_context():
    db.create_all()
    
    # Для локального запуску - створюємо тестового адміна
    if 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI'] and User.query.count() == 0:
        admin = User(username='admin', role='superadmin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print("✅ Створено тестового суперадміна: admin / admin123")

if __name__ == "__main__":
    print("\n" + "="*60)
    print("🚀 Flask додаток запущено!")
    
    if 'postgresql' in app.config['SQLALCHEMY_DATABASE_URI']:
        print("🌐 Режим: СЕРВЕР (PostgreSQL)")
    else:
        print("💻 Режим: ЛОКАЛЬНО (SQLite)")
        print("👤 Тестовий адмін: admin / admin123")
    
    print("📍 URL: http://localhost:5000")
    print("="*60 + "\n")
    
    app.run(debug=True)